#!/usr/bin/env python3
"""
Acceptance test for generated hot cost sheets.

Compares the BUDGET / DAY figure the generator produced against the one the
production accountant put in the real workbook, person by person, on both a
prep day and a shoot day.

The prep column is the point. A generator that carries one day cost per person
will match on shoot days and be wrong on every prep day, and the totals will
still look plausible — so scoring only shoot days would hide the defect.

    python verify_generated_hotcost.py hotcost.xlsx REAL_HOTCOST.xls \\
        --generated-shoot 102017 --real-shoot 102017 \\
        --generated-prep 101317 --real-prep "101317 PRESHOOT"
"""

from __future__ import annotations

import argparse
import sys

try:
    import openpyxl
    import xlrd
except ImportError:  # pragma: no cover
    sys.exit("openpyxl and xlrd are required")

TOLERANCE = 0.02
GEN_NAME, GEN_POS, GEN_BUDGET = 2, 3, 17     # 1-indexed, openpyxl
REAL_NAME, REAL_POS, REAL_BUDGET = 1, 2, 16  # 0-indexed, xlrd


def key(name: str) -> str:
    return "".join(ch for ch in (name or "").upper() if ch.isalnum())


def read_generated(path: str, sheet: str) -> dict[str, tuple[str, float]]:
    wb = openpyxl.load_workbook(path)
    match = next((n for n in wb.sheetnames if n.startswith(sheet)), None)
    if match is None:
        raise SystemExit(f"{sheet!r} not among {wb.sheetnames[:6]}…")
    ws = wb[match]
    out = {}
    for row in range(1, ws.max_row + 1):
        name = ws.cell(row, GEN_NAME).value
        budget = ws.cell(row, GEN_BUDGET).value
        if isinstance(name, str) and name.strip() and isinstance(budget, (int, float)):
            out[key(name)] = (str(ws.cell(row, GEN_POS).value or ""), float(budget))
    return out


def read_real(path: str, sheet: str) -> dict[str, tuple[str, float]]:
    book = xlrd.open_workbook(path)
    match = next((n for n in book.sheet_names() if n.strip() == sheet.strip()), None)
    if match is None:
        raise SystemExit(f"{sheet!r} not among {book.sheet_names()[:6]}…")
    ws = book.sheet_by_name(match)
    out = {}
    for row in range(ws.nrows):
        name = ws.cell_value(row, REAL_NAME)
        budget = ws.cell_value(row, REAL_BUDGET)
        if isinstance(name, str) and name.strip() and isinstance(budget, float) and budget:
            out[key(name)] = (str(ws.cell_value(row, REAL_POS)).strip(), float(budget))
    return out


def score(generated: dict, real: dict, label: str) -> tuple[int, int, list[str]]:
    ok = total = 0
    misses = []
    for person, (position, actual) in sorted(real.items()):
        if person not in generated:
            continue
        total += 1
        produced = generated[person][1]
        if abs(produced - actual) < TOLERANCE:
            ok += 1
        else:
            misses.append(f"{position[:22]:<24}generated {produced:>9,.2f}   "
                          f"accountant {actual:>9,.2f}")
    print(f"\n{label}")
    print("-" * 64)
    print(f"matched {ok} of {total}" + (f"  ({ok / total * 100:.0f}%)" if total else ""))
    for miss in misses[:8]:
        print(f"   {miss}")
    if len(misses) > 8:
        print(f"   … and {len(misses) - 8} more")
    return ok, total, misses


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("generated_xlsx")
    ap.add_argument("real_xls")
    ap.add_argument("--generated-shoot", default="102017")
    ap.add_argument("--real-shoot", default="102017")
    ap.add_argument("--generated-prep", default="101317")
    ap.add_argument("--real-prep", default="101317 PRESHOOT")
    args = ap.parse_args(argv)

    shoot_ok, shoot_n, _ = score(
        read_generated(args.generated_xlsx, args.generated_shoot),
        read_real(args.real_xls, args.real_shoot),
        "SHOOT DAY — budgeted day cost")
    prep_ok, prep_n, _ = score(
        read_generated(args.generated_xlsx, args.generated_prep),
        read_real(args.real_xls, args.real_prep),
        "PREP DAY — budgeted day cost")

    total_ok, total_n = shoot_ok + prep_ok, shoot_n + prep_n
    print(f"\n{'=' * 64}")
    print(f"overall {total_ok} of {total_n}"
          + (f"  ({total_ok / total_n * 100:.0f}%)" if total_n else ""))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
