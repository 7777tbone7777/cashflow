#!/usr/bin/env python3
"""
Ingest union rate cards into a normalised scale-rate repository.

Deliberately parses **rate cards, not collective bargaining agreements**. A CBA
is a couple of hundred pages of legal prose; the rate schedule attached to it is
a table. The table is where the numbers live, and it is the only part a budget
tool needs.

    python ingest_rate_cards.py "Guild Agreements/" -o rates.json

What a rate card supplies that a budget cannot:

  · **Scale**, so a budgeted rate can be checked against the minimum. Below-scale
    is a liability, not a saving.
  · **Effective dates.** IATSE, DGA and SAG scales step up on contract
    anniversaries. A schedule spanning an increase is wrong without them.
  · **Overtime rules, stated rather than inferred.** The Local 600 card says
    "1-1/2 after 8", "Minimum Call - 8 Hours", "43.2 hour guarantee". The hot
    cost generator currently derives those empirically from whatever a budget
    happens to state, which leaves it blind to any local the budget does not
    quantify.

Rates are stored as derived tables, never as copies of the source documents,
which are licensed.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Any, Iterable

try:
    import openpyxl
    import pdfplumber
except ImportError:  # pragma: no cover
    sys.exit("openpyxl and pdfplumber are required")

MONEY = re.compile(r"\$?\s?([\d,]+\.\d{2}|[\d,]{3,})")
# A scale rate is never trivially small and never seven figures.
PLAUSIBLE = (5.0, 100_000.0)

LOCAL_PATTERNS = [
    (re.compile(r"\bLocal\s*#?\s*(\d{2,4})\b", re.I), "IATSE Local {0}"),
    (re.compile(r"\bIATSE\s+LOCAL\s+(\d{2,4})\b", re.I), "IATSE Local {0}"),
    (re.compile(r"\bDIRECTORS\s+GUILD\b|\bDGA\b", re.I), "DGA"),
    (re.compile(r"\bWRITERS\s+GUILD\b|\bWGA\b", re.I), "WGA"),
    (re.compile(r"\bSAG-?AFTRA\b", re.I), "SAG-AFTRA"),
    (re.compile(r"\bTeamsters?\b|\bLocal\s*399\b", re.I), "Teamsters Local 399"),
]

DATE_RANGE = re.compile(
    r"(?:effective[:\s]*)?"
    r"([A-Z][a-z]+\s+\d{1,2},?\s+\d{4})\s*(?:to|-|through|–)\s*"
    r"([A-Z][a-z]+\s+\d{1,2},?\s+\d{4})", re.I)
DATE_SLASH = re.compile(r"(\d{1,2}/\d{1,2}/\d{2,4})\s*-\s*(\d{1,2}/\d{1,2}/\d{2,4})")

# Overtime and guarantee language, as printed on the cards.
OT_AFTER = re.compile(r"1-?1/2\s*(?:x\s*)?after\s*([\d.]+)", re.I)
MIN_CALL = re.compile(r"Minimum\s*Call\s*-?\s*([\d.]+)\s*Hours?", re.I)
GUARANTEE = re.compile(r"([\d.]+)\s*hours?\s*;?\s*$|weekly\s*guarantee", re.I)

PRODUCTION_TYPE_HINTS = [
    (re.compile(r"\bfeature", re.I), "feature"),
    (re.compile(r"\bpilot|episodic|season|television|\bTV\b|SVOD", re.I), "television"),
]


@dataclass
class ScaleRate:
    occupation_code: str | None
    classification: str
    basis: str                 # hour | day | week
    amount: float
    production_type: str | None = None   # feature | television | None = both


@dataclass
class RateCard:
    source: str
    local: str | None = None
    region: str | None = None
    panel: str | None = None
    effective_from: str | None = None
    effective_to: str | None = None
    minimum_call_hours: float | None = None
    overtime_after_hours: float | None = None
    weekly_guarantee_hours: float | None = None
    rates: list[ScaleRate] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    confidence: float = 1.0
    candidate_lines: int = 0
    trusted: bool = True


# Prose is full of numbers that are not money. Years are the worst offender:
# "the 2023 Agreement" reads as $2,023.00 to any naive money regex.
YEARS = {float(y) for y in range(1985, 2101)}

# Addresses, URLs and boilerplate read as title-case job titles to a naive
# filter. The Local 600 card yields "70 W 36th Street 9th Floor" at $877/hour
# without these.
NOT_AN_OCCUPATION = re.compile(
    r"www\.|https?:|@|\.com|\.org|\bstreet\b|\bavenue\b|\bfloor\b|\bsuite\b|"
    r"\bp\.?o\.?\s*box\b|\beffective\b|\brevised\b|\bpage\b|\btotal\b|"
    r"\bagreement\b|\bschedule\b|\bexhibit\b|\bappendix\b|\breport your\b|"
    r"\btelephone\b|\bfax\b|\bemail\b", re.I)

STOPWORD_TAIL = re.compile(
    r"\b(the|of|to|as|and|or|a|an|in|on|for|with|shall|be|is|are|was|were|"
    r"under|per|from|by|that|this|which|article|section|paragraph)\s*$", re.I)
SENTENCE_ISH = re.compile(r"[,;:]\s|\.\s|“|”|\(\s*[a-z]\s*\)")


def to_money(token: str, *, had_dollar: bool = False) -> float | None:
    cleaned = token.replace("$", "").replace(",", "").strip()
    try:
        value = float(cleaned)
    except ValueError:
        return None
    if not PLAUSIBLE[0] <= value <= PLAUSIBLE[1]:
        return None
    # A bare four-digit integer in running text is almost always a year.
    if not had_dollar and value in YEARS and float(int(value)) == value:
        return None
    return value


def looks_like_classification(name: str) -> bool:
    """Is this a job title, or a fragment of a sentence?

    Rate cards list occupations. Agreements discuss them in prose. Without this
    check a 200-page MBA yields dozens of confident-looking rates built from
    clause fragments and section numbers.
    """
    text = (name or "").strip()
    if not 3 < len(text) <= 60:
        return False
    if STOPWORD_TAIL.search(text) or SENTENCE_ISH.search(text):
        return False
    if NOT_AN_OCCUPATION.search(text):
        return False
    if re.match(r"^\d{1,4}\s+[A-Z]", text):        # street numbers
        return False
    words = text.split()
    if not 1 <= len(words) <= 8:
        return False
    letters = sum(ch.isalpha() for ch in text)
    return letters >= max(3, len(text) * 0.5)


def detect_local(text: str) -> str | None:
    for pattern, template in LOCAL_PATTERNS:
        m = pattern.search(text)
        if m:
            return template.format(*m.groups()) if m.groups() else template
    return None


def detect_dates(text: str) -> tuple[str | None, str | None]:
    m = DATE_RANGE.search(text) or DATE_SLASH.search(text)
    return (m.group(1).strip(), m.group(2).strip()) if m else (None, None)


def detect_conditions(text: str) -> dict[str, float | None]:
    ot = OT_AFTER.search(text)
    call = MIN_CALL.search(text)
    guarantee = re.search(r"([\d.]{2,5})\s*hours?\s*;", text)
    return {
        "overtime_after_hours": float(ot.group(1)) if ot else None,
        "minimum_call_hours": float(call.group(1)) if call else None,
        "weekly_guarantee_hours": float(guarantee.group(1)) if guarantee else None,
    }


def classify_columns(header_cells: list[str]) -> list[tuple[int, str, str | None]]:
    """Map each money column to (index, basis, production_type)."""
    out = []
    for i, cell in enumerate(header_cells):
        text = (cell or "").lower()
        basis = ("week" if "week" in text else
                 "day" if "day" in text else
                 "hour" if "hour" in text or "hr" in text else None)
        production = None
        for pattern, name in PRODUCTION_TYPE_HINTS:
            if pattern.search(cell or ""):
                production = name
                break
        if basis or production:
            out.append((i, basis or "hour", production))
    return out


def parse_xlsx(path: Path) -> RateCard | None:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    grid = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]
    blob = " ".join(str(v) for row in grid for v in row if v)

    card = RateCard(source=path.name, local=detect_local(blob))
    card.effective_from, card.effective_to = detect_dates(blob)
    for key, value in detect_conditions(blob).items():
        setattr(card, key, value)
    region = re.search(r"([A-Z][a-z]+\s+Region)", blob)
    panel = re.search(r"(PANEL\s*\d+)", blob, re.I)
    card.region = region.group(1) if region else None
    card.panel = panel.group(1) if panel else None

    # The row naming the columns is the one mentioning an hour or a week.
    header_row = next((i for i, row in enumerate(grid)
                       if sum(1 for v in row if v and re.search(
                           r"per\s*(hour|day|week)", str(v), re.I)) >= 2), None)
    if header_row is None:
        return card
    headers = [str(v or "") for v in grid[header_row]]
    columns = classify_columns(headers)

    for row in grid[header_row + 1:]:
        cells = [str(v) if v is not None else "" for v in row]
        code = next((c for c in cells[:2] if re.fullmatch(r"\d{3,5}", c.strip())), None)
        name = next((c for c in cells if c and not re.fullmatch(
            r"[\d.,$\s]+", c) and looks_like_classification(c)), None)
        if not name:
            continue
        for index, basis, production in columns:
            if index >= len(cells):
                continue
            amount = to_money(cells[index])
            if amount is None:
                continue
            card.rates.append(ScaleRate(
                occupation_code=code, classification=name.strip(),
                basis=basis, amount=amount, production_type=production))
    return card


def parse_pdf(path: Path, max_pages: int = 12) -> RateCard | None:
    with pdfplumber.open(path) as pdf:
        pages = pdf.pages[:max_pages]
        text = "\n".join((p.extract_text() or "") for p in pages)
        card = RateCard(source=path.name, local=detect_local(text))
        card.effective_from, card.effective_to = detect_dates(text)
        for key, value in detect_conditions(text).items():
            setattr(card, key, value)
        region = re.search(r"([A-Z][a-z]+\s+Region)", text)
        card.region = region.group(1) if region else None

        # Column headers that split feature from television, where present.
        split = re.search(r"CATEGORY\s+(FEATURE)\s+(TELEVISION)", text, re.I)
        production_columns = ["feature", "television"] if split else [None]

        # Confidence is measured per page and the best page wins. A rate card
        # has at least one page that is mostly rate rows; a 200-page agreement
        # with a wage appendix has exactly one too, and that appendix is the part
        # worth keeping. Averaging over the whole document hides both.
        page_scores: list[float] = []
        for page in pages:
            page_text = page.extract_text() or ""
            lines = [l for l in page_text.split("\n") if MONEY.search(l)]
            if len(lines) < 4:
                continue
            good = sum(1 for l in lines
                       if looks_like_classification(
                           re.sub(r"\s{2,}", " ", MONEY.split(l)[0].strip(" .$\t"))))
            page_scores.append(good / len(lines))

        candidate_lines = 0
        for line in text.split("\n"):
            raw = MONEY.findall(line)
            if not raw:
                continue
            candidate_lines += 1
            had_dollar = "$" in line
            amounts = [to_money(m, had_dollar=had_dollar) for m in raw]
            amounts = [a for a in amounts if a is not None]
            if not amounts:
                continue
            name = MONEY.split(line)[0].strip(" .$\t")
            name = re.sub(r"\s{2,}", " ", name)
            if not looks_like_classification(name):
                continue
            basis = ("week" if re.search(r"weekly|per week", line, re.I) else
                     "day" if re.search(r"daily|per day", line, re.I) else "hour")
            code = re.match(r"^(\d{3,5})\b", name)
            for i, amount in enumerate(amounts[:len(production_columns)]):
                card.rates.append(ScaleRate(
                    occupation_code=code.group(1) if code else None,
                    classification=name,
                    basis=basis, amount=amount,
                    production_type=production_columns[i]
                    if i < len(production_columns) else None))

        # A rate card is mostly rate rows; an agreement that happens to mention
        # money is not. Score it and let the caller decide what to trust.
        card.candidate_lines = candidate_lines
        card.confidence = max(page_scores) if page_scores else 0.0
        return card


def ingest(paths: Iterable[Path]) -> tuple[list[RateCard], list[str]]:
    cards, skipped = [], []
    for path in sorted(paths):
        try:
            card = (parse_xlsx(path) if path.suffix.lower() in (".xlsx", ".xlsm")
                    else parse_pdf(path))
        except Exception as exc:                      # noqa: BLE001
            skipped.append(f"{path.name}: {type(exc).__name__} {exc}")
            continue
        if card is None or not card.rates:
            skipped.append(f"{path.name}: no rate table found")
            continue
        # Trust a card only when most of its money-bearing lines resolved to
        # occupations. Everything else is kept but flagged, never silently used.
        card.trusted = card.confidence >= 0.60 and len(card.rates) >= 8
        cards.append(card)
    return cards, skipped


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("source", help="a rate card, or a directory of them")
    ap.add_argument("-o", "--out", default="rates.json")
    ap.add_argument("--max-pages", type=int, default=12)
    args = ap.parse_args(argv)

    root = Path(args.source)
    paths = ([p for p in root.rglob("*") if p.suffix.lower() in
              (".pdf", ".xlsx", ".xlsm")] if root.is_dir() else [root])

    cards, skipped = ingest(paths)
    trusted = [c for c in cards if c.trusted]
    payload = {
        "card_count": len(cards),
        "trusted_card_count": len(trusted),
        "trusted_rate_count": sum(len(c.rates) for c in trusted),
        "rate_count": sum(len(c.rates) for c in cards),
        "cards": [asdict(c) for c in cards],
        "skipped": skipped,
        "_note": ("Derived scale tables. Source agreements are licensed and are "
                  "deliberately not reproduced here."),
    }
    Path(args.out).write_text(json.dumps(payload, indent=2))

    print(f"wrote {args.out}")
    print(f"\n{len(paths)} documents · {len(cards)} yielded a rate table · "
          f"{len(trusted)} trusted")
    print(f"{payload['trusted_rate_count']:,} trusted rates "
          f"({payload['rate_count']:,} total before confidence filtering)")
    print(f"{len(skipped)} documents had no parseable rate table")
    untrusted = [c for c in cards if not c.trusted]
    if untrusted:
        print(f"\n{len(untrusted)} flagged low-confidence and excluded from lookups:")
        for c in sorted(untrusted, key=lambda c: c.confidence)[:6]:
            print(f"   {c.source[:52]:<54}confidence {c.confidence:.0%}")

    print(f"\n{'LOCAL':<26}{'EFFECTIVE':<26}{'RATES':>7}  CONDITIONS")
    print("-" * 84)
    for card in sorted(trusted, key=lambda c: -len(c.rates))[:12]:
        window = (f"{card.effective_from or '?'} → {card.effective_to or '?'}"
                  )[:25]
        conditions = []
        if card.minimum_call_hours:
            conditions.append(f"call {card.minimum_call_hours:g}h")
        if card.overtime_after_hours:
            conditions.append(f"1.5x after {card.overtime_after_hours:g}h")
        if card.weekly_guarantee_hours:
            conditions.append(f"gtee {card.weekly_guarantee_hours:g}h")
        print(f"{(card.local or '—')[:25]:<26}{window:<26}"
              f"{len(card.rates):>7}  {', '.join(conditions)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
