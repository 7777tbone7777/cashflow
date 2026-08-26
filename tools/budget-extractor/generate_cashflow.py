#!/usr/bin/env python3
"""
Generate a weekly cash flow schedule and hot cost day sheets from an extracted
budget plus a small production config.

    python generate_cashflow.py budget.json production.json -o cashflow.json
    python generate_cashflow.py budget.json production.json --compare REAL.xlsx

Design notes worth knowing before changing anything here:

  · Money is placed in three tiers, best first. A phase line knows its own
    duration and phase, so it is spread over the weeks it actually spans. A
    fringe follows the wages it is charged on, shifted by its own remittance
    lag. Anything left — lump sums with no phase detail — falls back to a
    department archetype, and every dollar placed that way is reported, because
    an archetype is a guess and guesses should be visible.

  · Cost is converted to cash at the end, never during. Each account class is
    shifted by its own payment behaviour (payroll lag, vendor terms, deposits
    up front). Keeping that as a final pass means the cost curve stays
    inspectable and the cash curve is reproducible from it.

  · The total is asserted, not hoped for. If the grid does not sum to the
    budget, the tool says so rather than emitting a plausible-looking schedule.
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, timedelta
from typing import Any

from overrides import (AMEND, OverrideSet, target_keys)

DAYS_PER_WEEK = 5.0

# Fallback spreads, expressed as the share of a department's money landing in
# each phase. Derived from a real production's cash flow — see README. Used only
# where the budget gives no phase detail, and always reported.
ARCHETYPES: dict[str, dict[str, float]] = {
    "pure_post":      {"post": 1.00},
    "post_heavy":     {"prep": 0.01, "shoot": 0.18, "post": 0.81},
    "straddle":       {"shoot": 0.40, "post": 0.60},
    "shoot_only":     {"shoot": 1.00},
    "shoot_heavy":    {"prep": 0.15, "shoot": 0.83, "post": 0.02},
    "build_ahead":    {"prep": 0.45, "shoot": 0.51, "post": 0.04},
    "whole_show":     {"prep": 0.36, "shoot": 0.46, "post": 0.18},
    "prep_only":      {"prep": 1.00},
    "wrap_tail":      {"prep": 0.02, "shoot": 0.02, "post": 0.96},
    "front_loaded":   {"prep": 0.90, "post": 0.10},
}

DEPARTMENT_ARCHETYPE: dict[str, str] = {
    "1100": "front_loaded",  "1200": "whole_show",   "1300": "whole_show",
    "1400": "shoot_heavy",   "2000": "whole_show",   "2100": "shoot_only",
    "2200": "build_ahead",   "2300": "build_ahead",  "2400": "wrap_tail",
    "2500": "shoot_heavy",   "2600": "shoot_heavy",  "2700": "build_ahead",
    "2800": "build_ahead",   "2900": "build_ahead",  "3000": "shoot_heavy",
    "3100": "shoot_heavy",   "3200": "shoot_heavy",  "3300": "shoot_heavy",
    "3400": "shoot_heavy",   "3500": "shoot_heavy",  "3600": "shoot_heavy",
    "3700": "shoot_only",    "4000": "shoot_only",   "4100": "prep_only",
    "4200": "shoot_heavy",   "4400": "straddle",     "4500": "post_heavy",
    "4600": "pure_post",     "4700": "pure_post",    "4800": "pure_post",
    "4900": "pure_post",     "6500": "post_heavy",   "6700": "shoot_heavy",
    "6800": "wrap_tail",
}

# Post spend declines as the show runs out. An archetype says how much of an
# account lands in post; it said nothing about when, so every post account was
# spread level across every post week — a flat line for twenty weeks, which is
# the one thing post never looks like.
#
# It declines because the cutting rooms empty. The full crew is there when the
# picture is assembled and rolls off through the finish, so the weekly number
# falls even though the finishing crafts are still working. What does not
# decline is delivery: the money tied to it is scheduled as a milestone rather
# than smeared, which is a separate question the app already asks.
#
# The taper is the last post week as a share of the first, and 1.0 restores the
# flat line.
#
# The direction is not in doubt. The steepness is one observation: on The Mule
# the back nine weeks of post averaged 0.60 of the front seven, which a straight
# decline reproduces at about 0.33 — and 0.33 is also where the error over those
# weeks bottoms out. Setting the default there would be fitting the only show
# available and calling it accuracy, which is how this project got a correlation
# it had to withdraw. Half by the end is the honest reading of one data point:
# it takes most of the improvement (70,030 to 54,526 mean weekly error against
# The Mule, versus 50,168 at the fitted value) without pretending to a precision
# nothing supports. An accountant who knows their own show should change it.
DEFAULT_POST_TAPER = 0.5

# How each class of spend converts from cost incurred to cash paid.
DEFAULT_TIMING = {
    "labour":    {"lag_days": 7,  "note": "payroll funded the week after work"},
    "fringe":    {"lag_days": 21, "note": "statutory and union remittance"},
    "vendor":    {"lag_days": 30, "note": "net-30"},
    "prepaid":   {"lag_days": -14, "note": "paid before the period it covers"},
}
PREPAID_DEPARTMENTS = {"6700"}          # insurance premium
PREPAID_KEYWORDS = ("deposit", "bond", "premium", "insurance", "advance")


def parse_iso(value: str | None) -> date | None:
    """Tolerant: an unparsed start date is a missing anchor, not a crash."""
    try:
        return date.fromisoformat(value) if value else None
    except (ValueError, TypeError):
        return None


def week_ending(day: date, weekday: int = 5) -> date:
    """Saturday-ending weeks by default; matches how cash flows are cut."""
    return day + timedelta(days=(weekday - day.weekday()) % 7)


@dataclass
class Period:
    index: int
    week_ending: date
    label: str
    phase: str

    def as_dict(self) -> dict[str, Any]:
        return {"index": self.index, "week_ending": self.week_ending.isoformat(),
                "label": self.label, "phase": self.phase}


@dataclass
class Calendar:
    periods: list[Period]
    shoot_start: date
    shoot_end: date

    def index_for(self, day: date) -> int | None:
        target = week_ending(day)
        for p in self.periods:
            if p.week_ending == target:
                return p.index
        if self.periods and target < self.periods[0].week_ending:
            return 0
        if self.periods and target > self.periods[-1].week_ending:
            return len(self.periods) - 1
        return None

    def indices_for_phase(self, phase: str) -> list[int]:
        return [p.index for p in self.periods if p.phase == phase]


def build_calendar(cfg: dict[str, Any]) -> Calendar:
    shoot_start = parse_iso(cfg["shoot_start"])
    if shoot_start is None:
        raise SystemExit("production config needs a valid ISO shoot_start")
    shoot_weeks = cfg["shoot_weeks"]
    prep_weeks = cfg["prep_weeks"]
    wrap_weeks = cfg.get("wrap_weeks", 1)
    post_weeks = cfg.get("post_weeks", 0)
    hiatus_after = set(cfg.get("hiatus_after_post_week", []))

    first_shoot_we = week_ending(shoot_start)
    periods: list[Period] = []
    idx = 0

    for i in range(prep_weeks, 0, -1):
        we = first_shoot_we - timedelta(weeks=i)
        label = f"Prep Wk (-{i})" if i <= 2 else f"Pre-Prep ({-i})"
        periods.append(Period(idx, we, label, "prep"))
        idx += 1

    for i in range(shoot_weeks):
        periods.append(Period(idx, first_shoot_we + timedelta(weeks=i),
                              f"Shoot Wk {i + 1}", "shoot"))
        idx += 1

    after_shoot = first_shoot_we + timedelta(weeks=shoot_weeks)
    for i in range(wrap_weeks):
        periods.append(Period(idx, after_shoot + timedelta(weeks=i),
                              f"Wrap Wk {i + 1}", "wrap"))
        idx += 1

    cursor = after_shoot + timedelta(weeks=wrap_weeks)
    post_number = 0
    while post_number < post_weeks:
        post_number += 1
        periods.append(Period(idx, cursor, f"Post Wk {post_number}", "post"))
        idx += 1
        cursor += timedelta(weeks=1)
        if post_number in hiatus_after:
            for _ in range(cfg.get("hiatus_weeks", 0)):
                periods.append(Period(idx, cursor, "Hiatus", "hiatus"))
                idx += 1
                cursor += timedelta(weeks=1)

    shoot_end = first_shoot_we + timedelta(weeks=shoot_weeks - 1)
    return Calendar(periods, shoot_start, shoot_end)


@dataclass
class Placement:
    """One amount, placed into one week, with a record of how it got there."""
    period_index: int
    amount: float
    acct: str
    department: str
    description: str
    basis: str            # phase_line | fringe | archetype
    phase: str
    cash_class: str

    def as_dict(self) -> dict[str, Any]:
        return {"period": self.period_index, "amount": round(self.amount, 2),
                "acct": self.acct, "department": self.department,
                "description": self.description, "basis": self.basis,
                "phase": self.phase, "cash_class": self.cash_class}


class Generator:
    def __init__(self, budget: dict[str, Any], cfg: dict[str, Any],
                 archetypes: dict[str, Any] | None = None,
                 overrides: OverrideSet | None = None) -> None:
        self.budget = budget
        self.cfg = cfg
        self.calendar = build_calendar(cfg)
        self.placements: list[Placement] = []
        self.notes: list[str] = []
        self.by_basis: dict[str, float] = defaultdict(float)
        self.learned = archetypes or {}
        self.archetype_hits: dict[str, int] = defaultdict(int)
        self.overrides = overrides or OverrideSet()
        self.live_keys: set[str] = set()
        # Accounts whose stated total was placed whole, fringes included.
        self.absorbed_fringe_accounts: set[str] = set()
        # How much post spend the decline reshaped, so it can be reported as the
        # assumption it is rather than applied quietly.
        self.post_tapered = 0.0

    def _shape_for(self, acct: str, department: str) -> tuple[dict[str, float], str]:
        """Best available spread for an account, most specific first.

        A learned account shape beats a learned department shape, which beats the
        hand-written table. Which one was used is recorded, because a reviewer
        should be able to tell an observation from a default.
        """
        learned_accounts = self.learned.get("accounts", {})
        learned_departments = self.learned.get("departments", {})
        if acct in learned_accounts:
            self.archetype_hits["learned_account"] += 1
            return learned_accounts[acct]["shares"], f"learned:{acct}"
        if department in learned_departments:
            self.archetype_hits["learned_department"] += 1
            return learned_departments[department]["shares"], f"learned_dept:{department}"
        self.archetype_hits["hand_written"] += 1
        key = DEPARTMENT_ARCHETYPE.get(department, "shoot_heavy")
        return ARCHETYPES.get(key, {"shoot": 1.0}), f"default:{key}"

    # ---------- placement helpers -------------------------------------------

    def _post_weights(self, indices: list[int]) -> list[float] | None:
        """A straight line falling across the post window.

        Returns None when the taper is off or there is no real window to shape,
        so level remains the behaviour wherever a decline would be an assertion
        rather than a shape.
        """
        taper = self.cfg.get("post_taper", DEFAULT_POST_TAPER)
        try:
            taper = float(taper)
        except (TypeError, ValueError):
            return None
        if taper >= 1.0 or taper < 0.0 or len(indices) < 3:
            return None
        last = len(indices) - 1
        return [1.0 - (1.0 - taper) * (position / last) for position in range(len(indices))]

    def _spread(self, amount: float, indices: list[int], *, acct: str,
                department: str, description: str, basis: str, phase: str,
                cash_class: str, weights: list[float] | None = None) -> None:
        if not amount or not indices:
            return
        if weights is None and phase == "post":
            weights = self._post_weights(indices)
            if weights is not None:
                self.post_tapered += amount
        if weights is None:
            weights = [1.0] * len(indices)
        total_weight = sum(weights) or 1.0
        for index, weight in zip(indices, weights):
            self.placements.append(Placement(
                period_index=index, amount=amount * weight / total_weight,
                acct=acct, department=department, description=description,
                basis=basis, phase=phase, cash_class=cash_class))
        self.by_basis[basis] += amount

    def _phase_window(self, phase: str, detail: dict[str, Any],
                      line: dict[str, Any]) -> list[int]:
        """Which weeks a phase line occupies.

        Prep is anchored to the record's own hire date when the budget gives
        one, and otherwise counted back from the shoot start by the line's own
        duration — so a crane wanted a week early lands a week early, rather
        than being snapped to the shoot window.
        """
        cal = self.calendar
        weeks = max(1, int(round((line.get("days") or DAYS_PER_WEEK) / DAYS_PER_WEEK)))

        if phase == "shoot":
            return cal.indices_for_phase("shoot")
        if phase in ("post",):
            return cal.indices_for_phase("post") or cal.indices_for_phase("wrap")
        if phase in ("wrap", "hold", "other"):
            return (cal.indices_for_phase("wrap")
                    or cal.indices_for_phase("shoot")[-1:])
        if phase == "prep":
            lead = self.overrides.resolve(
                "prep_lead_weeks",
                target_keys(department=detail.get("_department"),
                            account=detail.get("_account")),
                default=None)
            if lead:
                weeks = max(weeks, int(lead))
            anchor = parse_iso(detail.get("start_date"))
            if anchor:
                begin = cal.index_for(anchor)
            else:
                first_shoot = cal.indices_for_phase("shoot")
                begin = (first_shoot[0] - weeks) if first_shoot else 0
            begin = max(0, begin if begin is not None else 0)
            prep_indices = cal.indices_for_phase("prep")
            window = list(range(begin, min(begin + weeks, len(cal.periods))))
            window = [i for i in window if i in prep_indices] or prep_indices[-weeks:]
            return window
        return []

    def _allowance_window(self, line: dict[str, Any],
                          department: str) -> list[int] | None:
        """Place an allowance over the span it states.

        This is where a crane rented a week before the shoot ends up in the right
        week. The budget says how long the thing is held; matching that span to
        the phase of the same length puts it where it belongs instead of
        collapsing it into the shoot.

        The span is centred in that phase rather than run from its first week.
        The budget states how long a thing is held and almost never states when
        it arrives, and taking the opening weeks turns that silence into a claim
        that everything lands on the first day. It does not: on the reference
        show it piled 2.3M into the first week of the shoot against 0.3M in the
        fifth, a staircase that is an artefact of the anchor and nothing else.
        Centring keeps the duration the budget gives and stops inventing a start
        date it does not.

        Returns None when the line states no duration at all ("1 Allow", "1 Fee"),
        which is a genuine unknown rather than something to fake.
        """
        days = line.get("days")
        if not days:
            return None
        weeks = max(1, int(round(days / DAYS_PER_WEEK)))
        cal = self.calendar
        windows = {phase: cal.indices_for_phase(phase)
                   for phase in ("prep", "shoot", "post")}
        windows = {k: v for k, v in windows.items() if v}
        if not windows:
            return None
        # Match the stated span to the phase closest to it in length.
        phase = min(windows, key=lambda k: abs(len(windows[k]) - weeks))
        window = windows[phase]
        if weeks >= len(window):
            # Longer than any single phase: run it forward from that phase.
            start = window[0]
            return list(range(start, min(start + weeks, len(cal.periods))))
        offset = (len(window) - weeks) // 2
        return window[offset:offset + weeks]

    def _profile_for(self, acct: str, department: str) -> tuple[dict[str, float], str] | None:
        """A learned week-by-week curve, indexed off the first shoot week.

        Phase shares alone flatten the thing that matters most: within prep,
        real spend ramps by two orders of magnitude from the first week to the
        last. The profile keeps that ramp.
        """
        for table, tag in ((self.learned.get("accounts", {}), "profile"),
                           (self.learned.get("departments", {}), "profile_dept")):
            entry = table.get(acct if tag == "profile" else department)
            if entry and entry.get("profile"):
                return entry["profile"], f"{tag}:{acct if tag == 'profile' else department}"
        return None

    def _place_by_profile(self, amount: float, profile: dict[str, float],
                          department: str, acct: str, description: str,
                          is_labour: bool) -> None:
        first_shoot = self.calendar.indices_for_phase("shoot")
        origin = first_shoot[0] if first_shoot else 0
        n = len(self.calendar.periods)
        weights: dict[int, float] = {}
        for offset, share in profile.items():
            index = origin + int(offset)
            if 0 <= index < n:
                weights[index] = weights.get(index, 0.0) + share
            elif index >= n:          # profile runs past this show's post window
                weights[n - 1] = weights.get(n - 1, 0.0) + share
            else:
                weights[0] = weights.get(0, 0.0) + share
        if not weights:
            return
        indices = sorted(weights)
        self._spread(amount, indices, acct=acct, department=department,
                     description=description, basis="learned_profile",
                     phase="mixed",
                     cash_class=self._cash_class(department, description, is_labour),
                     weights=[weights[i] for i in indices])

    def _place_by_archetype(self, amount: float, department: str, acct: str,
                            description: str, is_labour: bool) -> str:
        learned_profile = self._profile_for(acct, department)
        if learned_profile:
            profile, provenance = learned_profile
            self._place_by_profile(amount, profile, department, acct,
                                   description, is_labour)
            self.archetype_hits[provenance.split(":")[0]] += 1
            return provenance

        shape, provenance = self._shape_for(acct, department)
        usable = {phase: share for phase, share in shape.items()
                  if self.calendar.indices_for_phase(phase)}
        weight_total = sum(usable.values()) or 1.0
        basis = ("learned" if provenance.startswith("learned") else "archetype")
        for phase, share in usable.items():
            self._spread(amount * share / weight_total,
                         self.calendar.indices_for_phase(phase),
                         acct=acct, department=department, description=description,
                         basis=basis, phase=phase,
                         cash_class=self._cash_class(department, description,
                                                     is_labour))
        return provenance

    @staticmethod
    def _cash_class(department: str, description: str, is_labour: bool) -> str:
        text = description.lower()
        if department in PREPAID_DEPARTMENTS or any(k in text for k in PREPAID_KEYWORDS):
            return "prepaid"
        return "labour" if is_labour else "vendor"

    # ---------- the three tiers ---------------------------------------------

    def place_phase_lines(self) -> None:
        self.resolve_milestones()
        prefer_observed = self.cfg.get("prefer_learned_profiles", False)
        for account in self.budget["accounts"]:
            department = account["acct"][:2] + "00"

            # Register identity before any shortcut, so an override that targets
            # this account is never reported as orphaned.
            account_keys = target_keys(department=department,
                                       account=account["acct"])
            self.live_keys.update(account_keys.values())
            steered = any(o.key in account_keys.values()
                          for o in self.overrides.overrides)

            # Where a past production shows how this exact account was actually
            # spread, that observation beats anything inferred from durations.
            # But a human override beats both: a learned profile is what usually
            # happens, and the person typing knows what is happening on this show.
            if prefer_observed and account.get("total") and not steered:
                learned = self._profile_for(account["acct"], department)
                if learned:
                    profile, provenance = learned
                    # A dated milestone is knowledge the profile does not have.
                    # Carve it out first and shape only what is left.
                    carved = self._place_account_milestones(account, department)
                    profile, provenance = learned
                    self._place_by_profile(
                        account["total"] - carved, profile, department, account["acct"],
                        account["name_display"],
                        any(d.get("is_labour") for d in account["details"]))
                    self.archetype_hits[provenance.split(":")[0]] += 1
                    self.absorbed_fringe_accounts.add(account["acct"])
                    continue

            for detail in account["details"]:
                is_labour = detail.get("is_labour", False)
                detail["_department"] = department
                detail["_account"] = account["acct"]
                keys = target_keys(department=department, account=account["acct"],
                                   sub=detail.get("sub"), person=detail.get("person"))
                self.live_keys.update(keys.values())

                for line in detail["phases"]:
                    amount = line.get("amount") or 0.0
                    if not amount:
                        continue
                    milestone_at = self.milestone_index.get(id(line))
                    if milestone_at is not None:
                        self._spread(
                            amount, [milestone_at], acct=account["acct"],
                            department=department,
                            description=line.get("label_display")
                            or account["name_display"],
                            basis="milestone", phase="milestone",
                            cash_class=self._cash_class(
                                department, account["name_display"], is_labour))
                        continue
                    # A human can move a line to a different phase entirely —
                    # the crane wanted a week early sits here.
                    phase = self.overrides.resolve("phase_window", keys,
                                                   default=line["phase"])
                    basis = "phase_line"
                    if phase in ("allowance", "travel"):
                        indices = self._allowance_window(line, department)
                        if indices is None:
                            # A flat allowance states no duration at all, which
                            # is exactly the case an archetype exists for.
                            self._place_by_archetype(
                                amount, department, account["acct"],
                                account["name_display"], is_labour)
                            continue
                        basis = "allowance_duration"
                        # Rental houses and stages take a deposit before the
                        # gear ships. Without this the whole curve sits a week
                        # or two later than a real one.
                        deposit_share = self.cfg.get("rental_deposit_share", 0.0)
                        if deposit_share and not is_labour and len(indices) > 1:
                            deposit = amount * deposit_share
                            amount -= deposit
                            first = max(0, indices[0] - 1)
                            self._spread(
                                deposit, [first], acct=account["acct"],
                                department=department,
                                description=f"{account['name_display']} (deposit)",
                                basis="deposit", phase="prep",
                                cash_class="prepaid")
                    else:
                        indices = self._phase_window(phase, detail, line)
                    if not indices:
                        continue
                    self._spread(
                        amount, indices, acct=account["acct"], department=department,
                        description=detail.get("person")
                        or detail.get("sub_display") or account["name_display"],
                        basis=basis, phase=phase,
                        cash_class=self.overrides.resolve(
                            "cash_class", keys,
                            default=self._cash_class(
                                department, account["name_display"], is_labour)))

    def place_fringes(self) -> None:
        """Fringes ride the wages they are charged on, then remit on a lag."""
        wage_shape: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))
        for p in self.placements:
            if p.cash_class == "labour":
                wage_shape[p.department][p.period_index] += p.amount

        for account in self.budget["accounts"]:
            if not account.get("fringes"):
                continue
            if account["acct"] in self.absorbed_fringe_accounts:
                continue    # already inside the account total placed by profile
            department = account["acct"][:2] + "00"
            shape = wage_shape.get(department)
            for fringe in account["fringes"]:
                amount = fringe.get("amount") or 0.0
                if not amount:
                    continue
                if shape:
                    indices = list(shape.keys())
                    weights = [shape[i] for i in indices]
                else:
                    indices = self.calendar.indices_for_phase("shoot")
                    weights = None
                self._spread(amount, indices, acct=account["acct"],
                             department=department,
                             description=f"Fringe {fringe['code_display']}",
                             basis="fringe", phase="fringe",
                             cash_class="fringe", weights=weights)

    def place_remainder(self) -> None:
        """Close each department against its authoritative total.

        Reconciling per department rather than per account is deliberate. The
        department figure is the one Movie Magic prints and the top sheet is
        built from, so closing against it guarantees the grid sums to the
        budget. Whatever the budget did not itemise gets an archetype shape and
        is reported — an archetype is a guess, and a guess the reader cannot see
        is the thing that makes a schedule untrustworthy.
        """
        placed: dict[str, float] = defaultdict(float)
        for p in self.placements:
            placed[p.department] += p.amount

        # The top sheet is the authority on what a department holds, because it
        # is the statement that sums to the grand total. AccountTotalfor lines
        # only cover money that reached a detail page, and a percentage-derived
        # line ("6700 INSURANCE :1.3% $572,026") never does — it is printed on
        # the top sheet and nowhere else. Closing against AccountTotalfor alone
        # silently drops it.
        authoritative: dict[str, float] = {}
        for row in self.budget.get("topsheet") or []:
            # A department can appear on more than one top sheet row. Both rows
            # are real money in that department, so add rather than overwrite.
            authoritative[row["acct"]] = authoritative.get(row["acct"], 0.0) + row["total"]
        for acct, stated in (self.budget.get("department_totals") or {}).items():
            authoritative.setdefault(acct, stated)
        names = {r["acct"]: r["name_display"] for r in self.budget["topsheet"]}

        for department, stated in sorted(authoritative.items()):
            gap = stated - placed.get(department, 0.0)
            if abs(gap) < 1.0:
                continue
            if self._place_by_instalments(gap, department, names.get(department, department)):
                continue
            provenance = self._place_by_archetype(
                gap, department, department, names.get(department, department), False)
            share_of_dept = gap / stated if stated else 0.0
            self.notes.append(
                f"{department} {names.get(department, '')}: {gap:,.0f} "
                f"({share_of_dept * 100:.0f}% of the department) shaped by "
                f"{provenance} — the budget itemises no phase detail for it")

        self._close_to_grand_total()

    def _place_by_instalments(self, amount: float, department: str,
                              description: str) -> bool:
        """Place a top sheet line the detail pages never carried.

        An insurance premium or a bond fee is stated as a percentage of the
        budget and nowhere else. It has no phase, no duration and no date, so an
        archetype spreads it across the shoot — which for a premium billed as a
        deposit plus instalments is simply wrong. When someone who knows the
        policy states the instalments, use them.
        """
        schedule = (self.cfg.get("unbacked_line_schedule") or {}).get(department)
        if not schedule:
            return False
        total_share = sum(i.get("share", 0) for i in schedule) or 1.0
        placed_any = False
        for instalment in schedule:
            when = parse_iso(instalment.get("pay_on"))
            index = self.calendar.index_for(when) if when else None
            if index is None:
                continue
            self._spread(amount * instalment.get("share", 0) / total_share, [index],
                         acct=department, department=department,
                         description=description, basis="instalment",
                         phase="", cash_class="prepaid")
            placed_any = True
        if placed_any:
            self.notes.append(
                f"{department} {description}: {amount:,.0f} placed on the "
                f"{len(schedule)} instalment date(s) you supplied, not by archetype")
        return placed_any

    def _place_account_milestones(self, account: dict[str, Any],
                                  department: str) -> float:
        """Place this account's dated lines and report what they came to."""
        carved = 0.0
        for detail in account["details"]:
            for line in detail["phases"]:
                index = self.milestone_index.get(id(line))
                if index is None:
                    continue
                amount = line.get("amount") or 0.0
                self._spread(amount, [index], acct=account["acct"],
                             department=department,
                             description=line.get("label_display")
                             or account["name_display"],
                             basis="milestone", phase="milestone",
                             cash_class=self._cash_class(
                                 department, account["name_display"],
                                 any(d.get("is_labour") for d in account["details"])))
                carved += amount
        return carved

    def resolve_milestones(self) -> None:
        """Match each milestone rule to the budget lines it names.

        A bonus payable on delivery is priced by the budget and scheduled by
        nobody: every duration rule in this file spreads it across the shoot,
        where it lands months early. Only a person can say which week it pays.
        Resolving up front means the named lines are placed on their date and
        the rest of the account still flows through the normal path, so the
        total is untouched by construction.
        """
        self.milestone_index: dict[int, int] = {}       # id(phase line) -> period
        self.milestone_carved: dict[str, float] = defaultdict(float)
        for rule in self.cfg.get("milestones") or []:
            when = parse_iso(rule.get("pay_on"))
            index = self.calendar.index_for(when) if when else None
            label = f"{rule.get('acct') or ''} {rule.get('description') or ''}".strip()
            if index is None:
                self.notes.append(
                    f"milestone {label}: pay_on {rule.get('pay_on')!r} is not a week "
                    "in this calendar — left where the budget put it")
                continue
            acct = str(rule.get("acct") or "")
            needle = str(rule.get("description") or "").lower()
            if not acct and not needle:
                continue
            matched = 0.0
            for account in self.budget["accounts"]:
                if acct and account["acct"] != acct:
                    continue
                for detail in account["details"]:
                    for line in detail["phases"]:
                        if needle and needle not in str(
                                line.get("label_display") or "").lower():
                            continue
                        amount = line.get("amount") or 0.0
                        if not amount:
                            continue
                        self.milestone_index[id(line)] = index
                        self.milestone_carved[account["acct"]] += amount
                        matched += amount
            if matched:
                self.notes.append(
                    f"milestone {label}: {matched:,.0f} placed in the week ending "
                    f"{self.calendar.periods[index].week_ending} because you said so, "
                    "not by duration")
            else:
                self.notes.append(
                    f"milestone {label}: matched no line in this budget — check the "
                    "account number and the wording")

    def _with_post_note(self) -> list[str]:
        """Name the decline, with what it is worth, alongside every other guess."""
        if self.post_tapered <= 0:
            return self.notes
        taper = float(self.cfg.get("post_taper", DEFAULT_POST_TAPER))
        return self.notes + [
            f"post: {self.post_tapered:,.0f} was spread on a decline rather than level, "
            f"the last post week set to {taper:.0%} of the first — post empties out as the "
            "cutting rooms do. The direction is established; this steepness rests on one "
            "production, and it is a setting rather than a fact."
        ]

    def _close_to_grand_total(self) -> None:
        """Place whatever the departments still leave short of the stated total.

        Departments close against the top sheet, and the top sheet can still
        miss the grand total — rounding on a percentage line, or a row the
        print-out never carried. Emitting a grid that is short by a real amount
        is exactly the failure this generator exists to refuse, so the residual
        is placed and named rather than left to show up as a reconciliation
        error the reader cannot act on.
        """
        grand = self.budget.get("totals", {}).get("grand_total") or 0.0
        if not grand:
            return
        residual = grand - sum(p.amount for p in self.placements)
        if abs(residual) < 1.0:
            return

        # No department claims it, so the least-assumption shape is the one the
        # rest of the budget already has.
        weights = [0.0] * len(self.calendar.periods)
        for p in self.placements:
            weights[p.period_index] += p.amount
        indices = [i for i, w in enumerate(weights) if w > 0]
        if not indices:
            return
        self._spread(residual, indices, acct="", department="",
                     description="unattributed residual", basis="unattributed",
                     phase="", cash_class="vendor",
                     weights=[weights[i] for i in indices])

        if abs(residual) > grand * 0.0005:
            self.notes.append(
                f"unattributed: {residual:,.0f} ({residual / grand * 100:.2f}% of the "
                "budget) is stated in the grand total but claimed by no top sheet "
                "row — spread in proportion to everything else, and worth chasing")

    # ---------- cost -> cash --------------------------------------------------

    def to_cash(self) -> tuple[list[float], list[float]]:
        timing = {**DEFAULT_TIMING, **self.cfg.get("payment_timing", {})}
        # A department can settle on its own cycle — a payroll company on one
        # calendar, a location or transport vendor on another. Measured against
        # a real cash flow this is the highest-leverage answer in the config,
        # and the right value differs between productions, so it is asked and
        # never inferred.
        dept_lags = {str(k): float(v) for k, v in (timing.get("departments") or {}).items()}
        n = len(self.calendar.periods)
        cost = [0.0] * n
        cash = [0.0] * n
        for p in self.placements:
            cost[p.period_index] += p.amount
            lag_days = dept_lags.get(p.department,
                                     timing.get(p.cash_class, {}).get("lag_days", 0))
            # Split across the two weeks the lag straddles rather than rounding
            # to whole ones. Rounding threw away every answer that was not a
            # multiple of seven — funding payroll on the Wednesday for the week
            # just worked is four days, and it was indistinguishable from seven
            # — and it made the lag behave erratically, since a day either side
            # of the half-week could move a department's whole spend a week.
            # The defaults are exact weeks, so none of them move.
            weeks = lag_days / 7.0
            base = math.floor(weeks)
            carry = weeks - base
            for step, share in ((base, 1.0 - carry), (base + 1, carry)):
                if share <= 0:
                    continue
                target = min(max(p.period_index + step, 0), n - 1)
                cash[target] += p.amount * share
        return cost, cash

    # ---------- output --------------------------------------------------------

    def build(self) -> dict[str, Any]:
        self.place_phase_lines()
        placed_before_remainder = sum(p.amount for p in self.placements)
        self.place_fringes()
        self.place_remainder()

        # Redistribution may change when money lands, never how much. Assert it.
        budget_total = self.budget["totals"].get("grand_total") or 0.0
        self.overrides.find_orphans(self.live_keys)
        self.overrides.collect_amendments()
        if self.overrides.overrides:
            self.overrides.assert_total_preserved(
                budget_total, sum(p.amount for p in self.placements), tolerance=2.0)

        cost, cash = self.to_cash()

        by_department: dict[str, list[float]] = defaultdict(
            lambda: [0.0] * len(self.calendar.periods))
        for p in self.placements:
            by_department[p.department][p.period_index] += p.amount

        grand = self.budget["totals"].get("grand_total") or 0.0
        total_placed = sum(cost)
        cumulative, running = [], 0.0
        for value in cash:
            running += value
            cumulative.append(round(running, 2))

        return {
            "production": self.budget.get("production", {}),
            "config": self.cfg,
            "periods": [p.as_dict() for p in self.calendar.periods],
            "weekly_cost": [round(v, 2) for v in cost],
            "weekly_cash": [round(v, 2) for v in cash],
            "cumulative_cash": cumulative,
            "by_department": {k: [round(x, 2) for x in v]
                              for k, v in sorted(by_department.items())},
            "placement_basis": {k: round(v, 2) for k, v in self.by_basis.items()},
            "reconciliation": {
                "budget_grand_total": grand,
                "total_placed": round(total_placed, 2),
                "difference": round(total_placed - grand, 2),
                "reconciles": abs(total_placed - grand) < max(2.0, grand * 0.0005),
                "share_from_budget_detail": round(
                    (self.by_basis.get("phase_line", 0)
                     + self.by_basis.get("fringe", 0)) / total_placed, 4)
                if total_placed else 0.0,
            },
            "assumptions": self._with_post_note(),
            "archetype_provenance": dict(self.archetype_hits),
            "overrides": {
                "loaded": len(self.overrides.overrides),
                "applied": dict(self.overrides.applied),
                # Named, so a caller can tell which of them did anything.
                "applied_targets": sorted(self.overrides.applied_targets),
                "inert": sorted(
                    f"{o.field}|{o.scope}|{o.key}" for o in self.overrides.overrides
                    if f"{o.field}|{o.scope}|{o.key}" not in self.overrides.applied_targets
                    and o.key in {k for k in self.live_keys}),
                "orphaned": [o.key for o in self.overrides.unused],
                "amendments_not_applied": [o.key for o in self.overrides.amendments],
            },
        }


def compare_to_real(result: dict[str, Any], path: str) -> None:
    """Hold the generated weekly curve up against the accountant's own."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl needed for --compare", file=sys.stderr)
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    # Find the sheet by its landmark row, not by name — a workbook usually has a
    # summary tab whose name also contains "cash flow".
    sheet = row = None
    for name in wb.sheetnames:
        candidate = wb[name]
        found = next((r for r in range(1, candidate.max_row + 1)
                      if str(candidate.cell(r, 3).value or "")
                      .startswith("WEEKLY CASH FLOW")), None)
        if found:
            sheet, row = candidate, found
            break
    if sheet is None:
        print("no WEEKLY CASH FLOW TOTALS row found in any sheet", file=sys.stderr)
        return

    real: dict[date, float] = {}
    for col in range(7, 43):
        when = sheet.cell(2, col).value
        value = sheet.cell(row, col).value
        if hasattr(when, "date") and isinstance(value, (int, float)):
            real[when.date()] = float(value)

    print(f"\n{'WEEK END':<12}{'PERIOD':<16}{'GENERATED':>13}{'ACCOUNTANT':>13}{'DIFF':>13}")
    print("-" * 68)
    gen_total = real_total = 0.0
    for period, value in zip(result["periods"], result["weekly_cash"]):
        when = date.fromisoformat(period["week_ending"])
        actual = real.get(when)
        gen_total += value
        if actual is None:
            print(f"{str(when):<12}{period['label'][:15]:<16}{value:>13,.0f}"
                  f"{'—':>13}{'—':>13}")
            continue
        real_total += actual
        print(f"{str(when):<12}{period['label'][:15]:<16}{value:>13,.0f}"
              f"{actual:>13,.0f}{value - actual:>13,.0f}")
    print("-" * 68)
    print(f"{'TOTAL':<28}{gen_total:>13,.0f}{real_total:>13,.0f}"
          f"{gen_total - real_total:>13,.0f}")


def guard_production_type(budget: dict[str, Any], force: bool) -> None:
    """Refuse a television budget rather than quietly treating it as a feature.

    Television is episodic, with pattern and amortised costs and no single shoot
    block, and its detail pages nest subtotals differently — a feature-shaped
    schedule built from one would reconcile and still be meaningless. Detection
    exists precisely so this can fail loudly.
    """
    detected = (budget.get("production") or {}).get("production_type") or {}
    kind, confidence = detected.get("type"), detected.get("confidence", 0)
    if kind != "television":
        return
    if force:
        print(f"WARNING: this looks like a television budget "
              f"({confidence:.0%} confidence) and --force was given. The output "
              f"will not be trustworthy.\n")
        return
    reasons = "; ".join(detected.get("evidence", [])[:3])
    raise SystemExit(
        f"This appears to be a TELEVISION budget ({confidence:.0%} confidence: "
        f"{reasons}).\n\nTelevision is not supported yet — it is episodic, with "
        f"pattern and amortised costs and no single shoot block, so a "
        f"feature-shaped cash flow built from it would reconcile and still be "
        f"wrong.\n\nPass --force to override this check if the detection is "
        f"mistaken.")


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("budget_json")
    ap.add_argument("production_json")
    ap.add_argument("-o", "--out")
    ap.add_argument("--compare", help="a real cash flow xlsx to hold this against")
    ap.add_argument("--archetypes", help="learned spreads from learn_archetypes.py")
    ap.add_argument("--overrides", help="human corrections, applied over the defaults")
    ap.add_argument("--force", action="store_true",
                    help="generate even if the budget is not a feature")
    args = ap.parse_args(argv)

    budget = json.load(open(args.budget_json))
    cfg = json.load(open(args.production_json))
    guard_production_type(budget, args.force)

    archetypes = json.load(open(args.archetypes)) if args.archetypes else None
    generator = Generator(budget, cfg, archetypes,
                          OverrideSet.load(args.overrides))
    result = generator.build()

    if args.out:
        with open(args.out, "w") as fh:
            json.dump(result, fh, indent=2)
        print(f"wrote {args.out}", file=sys.stderr)

    rec = result["reconciliation"]
    print(f"\nperiods            {len(result['periods'])}")
    print(f"budget total       {rec['budget_grand_total']:>14,.2f}")
    print(f"placed             {rec['total_placed']:>14,.2f}")
    print(f"difference         {rec['difference']:>14,.2f}"
          f"   {'RECONCILES' if rec['reconciles'] else 'DOES NOT RECONCILE'}")
    print(f"from budget detail  {rec['share_from_budget_detail'] * 100:>13.1f}%"
          f"   (rest placed by archetype)")
    print("\nplacement basis:")
    for basis, amount in sorted(result["placement_basis"].items(),
                                key=lambda x: -x[1]):
        print(f"   {basis:<14}{amount:>14,.0f}")
    if result.get("archetype_provenance"):
        print("\nshapes used where the budget states no timing:")
        for source, count in sorted(result["archetype_provenance"].items(),
                                    key=lambda x: -x[1]):
            print(f"   {source:<22}{count:>5}")
    if result["assumptions"]:
        print(f"\nassumptions ({len(result['assumptions'])}) — every one is reviewable:")
        for note in result["assumptions"][:8]:
            print(f"   · {note}")

    generator.overrides.report()

    if args.compare:
        compare_to_real(result, args.compare)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
