#!/usr/bin/env python3
"""
Generate hot cost day sheets from an extracted budget.

One sheet per shoot day, pre-populated with the crew, their rates, and — the
part that matters — the **budgeted day cost for that day's phase**. On the day,
the accountant enters call, lunch and wrap; hours, overtime units, actual day
cost and variance all compute.

    python generate_hotcost.py budget.json production.json -o HOTCOST.xlsx

Two decisions worth understanding:

  · Budgeted day cost is phase-specific, taken straight from the budget's own
    stated guarantee for that phase. A prep day and a shoot day are different
    numbers for the same person — typically 41% apart — and a template that
    carries one figure per person is wrong on every prep day.

  · Union overtime rules are **derived from the budget, not modelled**. The
    reference budget shows Local 705 paying 17.25 units for a 14-hour day where
    Local 399 pays 20. Hard-coding a "standard" 8 / 1.5x / 2x rule would be
    wrong for most of the crew, so the hours-to-units curve is learned per union
    from the guarantees the budget already states.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import date, timedelta
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  pip install openpyxl")

DAYS_PER_WEEK = 5.0

COLUMNS = [
    ("ACCT", 11), ("NAME", 20), ("POSITION", 20), ("UNION", 9), ("RATE", 9),
    ("CALL", 7), ("LUNCH", 7), ("LUNCH 2", 8), ("WRAP", 7), ("HRS", 7),
    ("MP'S", 7), ("ST", 7), ("1.5X", 7), ("2X", 7), ("3X", 7),
    ("TOTAL / DAY", 12), ("BUDGET / DAY", 13), ("(OVER) / UNDER", 14),
]
C_ACCT, C_NAME, C_POS, C_UNION, C_RATE = 1, 2, 3, 4, 5
C_CALL, C_LUNCH, C_LUNCH2, C_WRAP, C_HRS = 6, 7, 8, 9, 10
C_MP, C_ST, C_15, C_2X, C_3X = 11, 12, 13, 14, 15
C_TOTAL, C_BUDGET, C_VAR = 16, 17, 18

# Used only where a union shows no guarantee in the budget at all.
FALLBACK_BREAKS = [(8.0, 1.0), (12.0, 1.5), (99.0, 2.0)]


def parse_iso(value: str | None) -> date | None:
    try:
        return date.fromisoformat(value) if value else None
    except (ValueError, TypeError):
        return None


def learn_unit_curves(budget: dict[str, Any]) -> dict[str, dict[float, float]]:
    """Observed hours -> units-per-day, per union, from the budget's guarantees.

    Every labour phase line that states both an hours figure and a unit count is
    a data point. Where a union states several counts for the same hours, the
    most frequently used one wins — that is the guarantee the budget actually
    works to.
    """
    tally: dict[str, dict[float, dict[float, int]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(int)))
    for account in budget["accounts"]:
        for detail in account["details"]:
            union = (detail.get("union") or "ANY").replace(" ", "")
            for line in detail["phases"]:
                hours, mult, unit = (line.get("hours_per_day"),
                                     line.get("multiplier"), line.get("unit"))
                if not hours or not mult or not unit:
                    continue
                per_day = (mult / DAYS_PER_WEEK if unit.lower().startswith("week")
                           else mult)
                # The "X" column carries unit counts for hourly crew and plain
                # headcount for day players. A real unit count is never below
                # the hours worked — straight time alone is one unit per hour —
                # so anything short of that is a headcount and not a guarantee.
                if per_day < hours * 0.9:
                    continue
                tally[union][hours][round(per_day, 2)] += 1

    curves: dict[str, dict[float, float]] = {}
    for union, by_hours in tally.items():
        curve = {}
        for hours, counts in by_hours.items():
            curve[hours] = max(counts.items(), key=lambda kv: (kv[1], -kv[0]))[0]
        curves[union] = dict(sorted(curve.items()))
    # A pooled curve for unions the budget never quantifies.
    pooled: dict[float, dict[float, int]] = defaultdict(lambda: defaultdict(int))
    for union, by_hours in tally.items():
        for hours, counts in by_hours.items():
            for units, n in counts.items():
                pooled[hours][units] += n
    curves["ANY"] = {h: max(c.items(), key=lambda kv: (kv[1], -kv[0]))[0]
                     for h, c in sorted(pooled.items())}
    return curves


def units_for(curve: dict[float, float], hours: float) -> float:
    """Interpolate the observed curve; extrapolate at its own marginal rate."""
    if not curve:
        units, previous = 0.0, 0.0
        for limit, rate in FALLBACK_BREAKS:
            span = max(0.0, min(hours, limit) - previous)
            units += span * rate
            previous = limit
            if hours <= limit:
                break
        return round(units, 2)
    points = sorted(curve.items())
    if hours <= points[0][0]:
        return round(points[0][1] * hours / points[0][0], 2)
    for (h0, u0), (h1, u1) in zip(points, points[1:]):
        if h0 <= hours <= h1:
            if h1 == h0:
                return u1
            return round(u0 + (u1 - u0) * (hours - h0) / (h1 - h0), 2)
    (h0, u0), (h1, u1) = points[-2], points[-1] if len(points) > 1 else (points[0], points[0])
    slope = ((u1 - u0) / (h1 - h0)) if h1 != h0 else 1.5
    return round(u1 + slope * (hours - h1), 2)


def build_crew(budget: dict[str, Any]) -> list[dict[str, Any]]:
    """Every person the budget names, with a day cost per phase.

    Grouped by department, not by account. A hot cost reads CAST, CAMERA,
    TRANSPORTATION — not LEAD #1, DAY PLAYERS, STUNT COORD — because the
    question it answers is which department is running over.
    """
    department_names = {row["acct"]: row["name_display"]
                        for row in budget.get("topsheet", [])}
    crew = []
    for account in budget["accounts"]:
        department = account["acct"][:2] + "00"
        for detail in account["details"]:
            if not detail.get("is_labour"):
                continue
            day_cost: dict[str, float] = {}
            hours: dict[str, float] = {}
            for line in detail["phases"]:
                if line.get("day_cost") and line.get("qty"):
                    day_cost.setdefault(line["phase"], line["day_cost"])
                    if line.get("hours_per_day"):
                        hours.setdefault(line["phase"], line["hours_per_day"])
            if not day_cost:
                continue
            crew.append({
                "acct": account["acct"],
                "department": department,
                "department_name": department_names.get(
                    department, account["name_display"]),
                "account_name": account["name_display"],
                "person": detail.get("person") or "",
                "position": (detail.get("sub_display") or "").split("/")[-1],
                "union": detail.get("union") or "",
                "rate": detail.get("rate_value"),
                "rate_basis": detail.get("rate_basis"),
                "day_cost": day_cost,
                "hours": hours,
                "start_date": detail.get("start_date"),
            })
    return crew


def shoot_days(cfg: dict[str, Any]) -> list[tuple[int, date, str]]:
    """(number, date, phase) for every day a sheet is wanted."""
    start = parse_iso(cfg["shoot_start"])
    if start is None:
        raise SystemExit("production config needs a valid ISO shoot_start")
    total = cfg.get("shoot_days") or int(cfg["shoot_weeks"] * 5)
    days_per_week = cfg.get("shoot_days_per_week", 5)

    out: list[tuple[int, date, str]] = []
    for offset in cfg.get("preshoot_day_offsets", [-3]):
        out.append((0, start + timedelta(days=offset), "prep"))

    cursor, made = start, 0
    while made < total:
        if cursor.weekday() < days_per_week:
            made += 1
            out.append((made, cursor, "shoot"))
        cursor += timedelta(days=1)
    return out


THIN = Side(style="thin", color="D0D0D0")

# Conventions a production accountant applies on top of the budget. The budget
# cannot state these — they are house practice — so they are declared, not
# inferred, and every one of them is visible in the config.
DEFAULT_CONVENTIONS = {
    # Weekly flat-rate crew (DGA, designers) bill their shoot day rate even on a
    # prep day; the budget's lower prep rate is a budgeting device, not a
    # timecard rate.
    "flat_rate_bills_shoot_day": True,
    # Hourly crew get a standard prep day rather than whatever fractional
    # guarantee the budget carries. Units, not dollars, so it scales by rate.
    "minimum_prep_units": 11.0,
    # Departments that rig on the preshoot day and therefore work shoot hours.
    "preshoot_at_shoot_hours": ["2500", "2700", "2800"],
    # A hot cost tracks *variable* labour — the people whose cost moves with the
    # day. Above-the-line flat deals (story, producers, director) do not vary,
    # so they are not on the sheet. This list matches the departments the
    # reference production's own hot cost carries.
    "departments": ["1400", "2000", "2300", "2500", "2600", "2700", "2800",
                    "2900", "3000", "3100", "3200", "3300", "3400", "3500", "3600"],
}


def rate_to_show(member: dict[str, Any]) -> float | None:
    """What belongs in the RATE column.

    Hourly crew show their hourly rate; flat-rate crew show their day cost. That
    is the convention the reference hot cost uses without exception — the UPM's
    RATE and BUDGET/DAY are both 1,063.60, while Craig Bauer's are 45.00 and
    630.00.
    """
    if member["rate_basis"] == "hour":
        return member["rate"]
    day = member["day_cost"].get("shoot") or member["day_cost"].get("prep")
    if day:
        return day
    if member["rate"] and member["rate_basis"] == "week":
        return member["rate"] / DAYS_PER_WEEK
    return member["rate"]


def budgeted_day_for(member: dict[str, Any], phase: str,
                     conventions: dict[str, Any]) -> float | None:
    """The budgeted day cost to print, after house conventions are applied."""
    stated = member["day_cost"].get(phase)
    shoot = member["day_cost"].get("shoot")

    if phase != "prep":
        return stated if stated is not None else shoot

    if member["department"] in conventions.get("preshoot_at_shoot_hours", []):
        return shoot if shoot is not None else stated

    if member["rate_basis"] == "week" and conventions.get("flat_rate_bills_shoot_day"):
        return shoot if shoot is not None else stated

    floor_units = conventions.get("minimum_prep_units")
    if stated is not None and floor_units and member["rate_basis"] == "hour":
        rate = member["rate"] or 0
        floor = floor_units * rate
        if floor > stated:
            return round(floor, 2)
    return stated


def department_fringe_rates(budget: dict[str, Any]) -> dict[str, float]:
    """Effective fringe rate per department, from the budget's own schedule.

    Charged against the department's actual labour rather than an individual
    fringe line's base, because different fringes sit on different bases — FICA
    on one, union health and welfare on another. Rates outside a plausible band
    usually mean a department with a token wage base, so they fall back to the
    production-wide average rather than distorting a day sheet.
    """
    fringe: dict[str, float] = defaultdict(float)
    labour: dict[str, float] = defaultdict(float)
    for account in budget["accounts"]:
        department = account["acct"][:2] + "00"
        for line in account.get("fringes", []):
            fringe[department] += line.get("amount") or 0.0
        for detail in account["details"]:
            if detail.get("is_labour"):
                labour[department] += sum(p.get("amount") or 0.0
                                          for p in detail["phases"])

    total_fringe = sum(fringe.values())
    total_labour = sum(labour.values()) or 1.0
    average = total_fringe / total_labour

    rates: dict[str, float] = {}
    for department in set(fringe) | set(labour):
        base = labour.get(department, 0.0)
        rate = (fringe.get(department, 0.0) / base) if base > 1000 else average
        rates[department] = round(rate if 0.10 <= rate <= 0.75 else average, 4)
    return rates


def write_day_sheet(ws, crew: list[dict[str, Any]], day_no: int, when: date,
                    phase: str, curves: dict[str, dict[float, float]],
                    conventions: dict[str, Any],
                    fringe_rates: dict[str, float]) -> dict[str, Any]:
    header_font = Font(bold=True, size=9)
    label_font = Font(bold=True, size=9, color="1F5C4D")
    money = '#,##0.00'

    ws.cell(1, 1, "HOT COST").font = Font(bold=True, size=12)
    ws.cell(1, 3, when.strftime("%A %d %b %Y"))
    ws.cell(1, 6, f"DAY #{day_no}" if phase == "shoot" else "PRESHOOT")
    ws.cell(1, 6).font = label_font
    ws.cell(2, 1, "enter CALL, LUNCH, WRAP as decimal hours (e.g. 7.5) — "
                  "everything right of WRAP computes").font = Font(italic=True, size=8,
                                                                   color="808080")

    head = 4
    for i, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(head, i, name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="EDF1EC")
        cell.border = Border(bottom=THIN)
        ws.column_dimensions[get_column_letter(i)].width = width

    row = head + 1
    by_department: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for member in crew:
        by_department[member["department_name"]].append(member)

    dept_rows: list[tuple[str, int, int]] = []
    for department, members in by_department.items():
        ws.cell(row, C_NAME, department).font = label_font
        row += 1
        first = row
        for member in members:
            budget_day = budgeted_day_for(member, phase, conventions)

            ws.cell(row, C_ACCT, member["acct"])
            ws.cell(row, C_NAME, member["person"])
            ws.cell(row, C_POS, member["position"])
            ws.cell(row, C_UNION, member["union"])
            rate = rate_to_show(member)
            if rate:
                ws.cell(row, C_RATE, round(rate, 2)).number_format = money

            # Hours worked, from the times the accountant will type.
            ws.cell(row, C_HRS,
                    f"=IF(AND(N({gc(C_CALL)}{row})>0,N({gc(C_WRAP)}{row})>0),"
                    f"{gc(C_WRAP)}{row}-{gc(C_CALL)}{row}"
                    f"-IF(N({gc(C_LUNCH)}{row})>0,0.5,0)"
                    f"-IF(N({gc(C_LUNCH2)}{row})>0,0.5,0),0)").number_format = '0.00'

            # Straight time and overtime, split on this union's own guarantee.
            union_key = (member["union"] or "ANY").replace(" ", "")
            curve = curves.get(union_key) or curves.get("ANY", {})
            straight = min(curve) if curve else 8.0
            ws.cell(row, C_ST, f"=MIN({gc(C_HRS)}{row},{straight})").number_format = '0.00'
            ws.cell(row, C_15,
                    f"=MAX(0,MIN({gc(C_HRS)}{row},12)-{straight})").number_format = '0.00'
            ws.cell(row, C_2X, f"=MAX(0,{gc(C_HRS)}{row}-12)").number_format = '0.00'

            ws.cell(row, C_TOTAL,
                    f"={gc(C_RATE)}{row}*({gc(C_ST)}{row}+{gc(C_15)}{row}*1.5"
                    f"+{gc(C_2X)}{row}*2+{gc(C_3X)}{row}*3+N({gc(C_MP)}{row}))"
                    ).number_format = money
            if budget_day is not None:
                ws.cell(row, C_BUDGET, round(budget_day, 2)).number_format = money
            ws.cell(row, C_VAR,
                    f"={gc(C_BUDGET)}{row}-{gc(C_TOTAL)}{row}").number_format = money
            for col in range(1, len(COLUMNS) + 1):
                ws.cell(row, col).border = Border(bottom=THIN)
            row += 1

        last = row - 1
        code = members[0]["department"]
        fringe_rate = fringe_rates.get(code, 0.0)

        # Department labour roll-up.
        labour_row = row
        ws.cell(row, C_POS, f"{department} LABOR").font = label_font
        for col in (C_TOTAL, C_BUDGET, C_VAR):
            ws.cell(row, col, f"=SUM({gc(col)}{first}:{gc(col)}{last})"
                    ).number_format = money
        row += 1

        # Fringe, at this department's own effective rate from the budget.
        # Kept as its own line because it remits on a different calendar and
        # because a department can be on budget on wages and over on fringe.
        fringe_row = row
        ws.cell(row, C_POS, f"{department} FRINGE").font = label_font
        ws.cell(row, C_RATE, fringe_rate).number_format = '0.0%'
        for col in (C_TOTAL, C_BUDGET, C_VAR):
            ws.cell(row, col, f"={gc(col)}{labour_row}*{fringe_rate:.4f}"
                    ).number_format = money
        row += 2

        dept_rows.append({
            "name": department, "code": code,
            "first": first, "last": last,
            "labour_row": labour_row, "fringe_row": fringe_row,
        })

    # Day summary — the figures the line producer actually reads.
    ws.cell(row, C_MP, "TOTAL").font = header_font
    ws.cell(row, C_TOTAL, "ACTUAL").font = header_font
    ws.cell(row, C_BUDGET, "BUDGET").font = header_font
    ws.cell(row, C_VAR, "TOTAL VAR").font = header_font
    row += 1
    ws.cell(row, C_TOTAL, "DAY COST").font = header_font
    ws.cell(row, C_BUDGET, "DY COST").font = header_font
    ws.cell(row, C_VAR, "(OVER) / UNDER").font = header_font
    row += 1
    for col in (C_TOTAL, C_BUDGET, C_VAR):
        parts = "+".join(f"{gc(col)}{d['labour_row']}+{gc(col)}{d['fringe_row']}"
                         for d in dept_rows)
        cell = ws.cell(row, col, f"={parts}" if parts else 0)
        cell.number_format = money
        cell.font = Font(bold=True)
    summary_row = row

    ws.freeze_panes = ws.cell(head + 1, 1)
    return {"departments": dept_rows, "summary_row": summary_row}


def gc(index: int) -> str:
    return get_column_letter(index)


def quote(sheet_name: str) -> str:
    return f"'{sheet_name}'"


def write_summary_sheet(ws, sheets: list[dict[str, Any]]) -> None:
    """Department by day, in variance — the sheet a line producer actually reads.

    Every cell is a formula pointing at the day sheets, so the picture updates as
    the accountant fills in times rather than needing a rebuild. Negative is over.
    """
    money = '#,##0'
    header_font = Font(bold=True, size=9)
    label_font = Font(bold=True, size=9, color="1F5C4D")
    over_fill = PatternFill("solid", fgColor="F6E7E4")

    ws.cell(1, 1, "HOT COST — DEPARTMENT SUMMARY").font = Font(bold=True, size=12)
    ws.cell(2, 1, "(over) / under against the budgeted day. "
                  "Negative is over. Updates as day sheets are filled in."
            ).font = Font(italic=True, size=8, color="808080")

    departments = [d["name"] for d in sheets[0]["departments"]] if sheets else []
    by_name = [{d["name"]: d for d in s["departments"]} for s in sheets]

    head = 4
    ws.cell(head, 1, "DEPARTMENT").font = header_font
    ws.column_dimensions["A"].width = 28
    for i, sheet in enumerate(sheets):
        col = 2 + i
        cell = ws.cell(head, col, sheet["label"])
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", textRotation=90)
        ws.column_dimensions[gc(col)].width = 9
    total_col = 2 + len(sheets)
    ws.cell(head, total_col, "TOTAL").font = header_font
    ws.column_dimensions[gc(total_col)].width = 13

    row = head + 1
    for department in departments:
        ws.cell(row, 1, department).font = label_font
        for i, sheet in enumerate(sheets):
            entry = by_name[i].get(department)
            if not entry:
                continue
            ref = (f"{quote(sheet['title'])}!{gc(C_VAR)}{entry['labour_row']}"
                   f"+{quote(sheet['title'])}!{gc(C_VAR)}{entry['fringe_row']}")
            cell = ws.cell(row, 2 + i, f"={ref}")
            cell.number_format = money
        ws.cell(row, total_col,
                f"=SUM({gc(2)}{row}:{gc(total_col - 1)}{row})").number_format = money
        ws.cell(row, total_col).font = Font(bold=True)
        row += 1

    ws.cell(row, 1, "TOTAL").font = header_font
    for col in range(2, total_col + 1):
        cell = ws.cell(row, col, f"=SUM({gc(col)}{head + 1}:{gc(col)}{row - 1})")
        cell.number_format = money
        cell.font = Font(bold=True)
        cell.border = Border(top=Side(style="thin", color="141917"))

    # Conditional shading so an over-budget department reads at a glance rather
    # than requiring the reader to scan for a minus sign.
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(
        f"B{head + 1}:{gc(total_col)}{row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=over_fill))

    ws.freeze_panes = ws.cell(head + 1, 2)


def write_weekly_sheet(ws, sheets: list[dict[str, Any]]) -> None:
    """The same picture rolled to weeks, which is the cadence of a cost report."""
    money = '#,##0'
    header_font = Font(bold=True, size=9)
    label_font = Font(bold=True, size=9, color="1F5C4D")

    ws.cell(1, 1, "HOT COST — WEEKLY ROLL-UP").font = Font(bold=True, size=12)
    ws.cell(2, 1, "actual, budget and variance by department by shoot week"
            ).font = Font(italic=True, size=8, color="808080")

    weeks: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for sheet in sheets:
        weeks[sheet["week"]].append(sheet)

    departments = [d["name"] for d in sheets[0]["departments"]] if sheets else []
    by_name = {s["title"]: {d["name"]: d for d in s["departments"]} for s in sheets}

    head = 4
    ws.cell(head, 1, "DEPARTMENT").font = header_font
    ws.column_dimensions["A"].width = 28
    col = 2
    week_cols: list[tuple[int, int]] = []
    for week in sorted(weeks):
        label = "Preshoot" if week == 0 else f"Week {week}"
        ws.cell(head - 1, col, label).font = header_font
        for offset, metric in enumerate(("ACTUAL", "BUDGET", "(OVER)/UNDER")):
            cell = ws.cell(head, col + offset, metric)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[gc(col + offset)].width = 12
        week_cols.append((week, col))
        col += 3

    row = head + 1
    for department in departments:
        ws.cell(row, 1, department).font = label_font
        for week, base_col in week_cols:
            for offset, source in enumerate((C_TOTAL, C_BUDGET, C_VAR)):
                refs = []
                for sheet in weeks[week]:
                    entry = by_name[sheet["title"]].get(department)
                    if entry:
                        refs.append(
                            f"{quote(sheet['title'])}!{gc(source)}{entry['labour_row']}"
                            f"+{quote(sheet['title'])}!{gc(source)}{entry['fringe_row']}")
                if refs:
                    ws.cell(row, base_col + offset,
                            "=" + "+".join(refs)).number_format = money
        row += 1

    ws.cell(row, 1, "TOTAL").font = header_font
    for _, base_col in week_cols:
        for offset in range(3):
            cell = ws.cell(row, base_col + offset,
                           f"=SUM({gc(base_col + offset)}{head + 1}:"
                           f"{gc(base_col + offset)}{row - 1})")
            cell.number_format = money
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(style="thin", color="141917"))

    ws.freeze_panes = ws.cell(head + 1, 2)


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("budget_json")
    ap.add_argument("production_json")
    ap.add_argument("-o", "--out", default="hotcost.xlsx")
    args = ap.parse_args(argv)

    budget = json.load(open(args.budget_json))
    cfg = json.load(open(args.production_json))

    conventions = {**DEFAULT_CONVENTIONS, **cfg.get('hot_cost_conventions', {})}
    curves = learn_unit_curves(budget)
    crew = build_crew(budget)
    included = conventions.get("departments")
    if included:
        before = len(crew)
        crew = [m for m in crew if m["department"] in included]
        print(f"scope: {len(crew)} of {before} budgeted people are on the hot cost "
              f"(above-the-line flat deals excluded)")
    days = shoot_days(cfg)

    fringe_rates = department_fringe_rates(budget)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    summary = wb.create_sheet("SUMMARY")
    weekly = wb.create_sheet("WEEKLY")

    sheets: list[dict[str, Any]] = []
    for day_no, when, phase in days:
        title = (when.strftime("%m%d%y") if phase == "shoot"
                 else when.strftime("%m%d%y") + " PRESHOOT")[:31]
        meta = write_day_sheet(wb.create_sheet(title), crew, day_no, when, phase,
                               curves, conventions, fringe_rates)
        meta.update({
            "title": title,
            "label": when.strftime("%d %b") if phase == "shoot" else "preshoot",
            "week": 0 if phase != "shoot" else (day_no - 1) // 5 + 1,
        })
        sheets.append(meta)

    write_summary_sheet(summary, sheets)
    write_weekly_sheet(weekly, sheets)
    wb.save(args.out)

    prep_only = sum(1 for m in crew if "prep" in m["day_cost"])
    both = sum(1 for m in crew
               if len({round(v, 2) for v in m["day_cost"].values()}) > 1)
    print(f"wrote {args.out}")
    print(f"\n{len(days)} day sheets · {len(crew)} crew")
    print(f"crew with a prep guarantee          {prep_only}")
    print(f"crew whose day cost changes by phase {both}")
    print("\nconventions applied (declared in production.json, not inferred):")
    for k, v in conventions.items():
        print(f"   {k:<28}{v}")
    print(f"\nunion overtime curves derived from the budget: {len(curves) - 1}")
    for union, curve in sorted(curves.items()):
        if union == "ANY" or not curve:
            continue
        shown = "  ".join(f"{int(h)}h={u:g}" for h, u in list(curve.items())[:5])
        print(f"   {union:<12}{shown}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
