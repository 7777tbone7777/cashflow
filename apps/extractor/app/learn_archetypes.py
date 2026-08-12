#!/usr/bin/env python3
"""
Learn account-level spread archetypes from cash flows that were cut by hand.

Hand-writing a table of "which weeks does box rental land in" is guesswork.
A completed cash flow is a labelled answer: every line item carries an account
code and its actual week-by-week allocation. This reads those allocations and
emits the observed phase shares per account, which the generator then uses in
place of a coarse departmental average.

Feed it as many past productions as you have. More shows means better priors,
and unlike a hand-written table it compounds.

    python learn_archetypes.py CASHFLOW.xlsx [MORE.xlsx ...] -o archetypes.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  pip install openpyxl")

PHASE_FROM_LABEL = [
    (re.compile(r"pre-?prep|prep", re.I), "prep"),
    (re.compile(r"shoot", re.I), "shoot"),
    (re.compile(r"hiatus", re.I), "hiatus"),
    (re.compile(r"post|wrap", re.I), "post"),
]

# Below this, a share is noise from a rounding artefact rather than a pattern.
MIN_SHARE = 0.005
# An account needs this much money before its shape means anything.
MIN_ACCOUNT_TOTAL = 500.0


def phase_for(label: str) -> str:
    for pattern, phase in PHASE_FROM_LABEL:
        if pattern.search(label or ""):
            return phase
    return "post"


def find_detail_sheet(wb: "openpyxl.Workbook"):
    """Locate the grid by its landmark row, not by sheet name."""
    for name in wb.sheetnames:
        ws = wb[name]
        for row in range(1, ws.max_row + 1):
            if str(ws.cell(row, 3).value or "").startswith("WEEKLY CASH FLOW"):
                return ws, row
    return None, None


def read_periods(ws, header_row: int) -> list[tuple[int, str, int]]:
    """Return (column, phase, offset) per period.

    `offset` is the week's position relative to the first shoot week — the only
    axis on which two different productions are comparable. Prep is negative,
    shoot starts at 0.
    """
    raw: list[tuple[int, str]] = []
    for col in range(7, ws.max_column + 1):
        label = ws.cell(header_row, col).value
        if label and str(label).strip():
            raw.append((col, phase_for(str(label))))
            continue
        when = ws.cell(header_row - 1, col).value
        if isinstance(when, datetime):
            raw.append((col, "post"))
    first_shoot = next((i for i, (_, phase) in enumerate(raw) if phase == "shoot"), 0)
    return [(col, phase, i - first_shoot) for i, (col, phase) in enumerate(raw)]


def harvest(path: str) -> dict[str, dict[str, float]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws, totals_row = find_detail_sheet(wb)
    if ws is None:
        print(f"  {path}: no cash flow grid found", file=sys.stderr)
        return {}

    header_row = next(
        (r for r in range(1, totals_row)
         if any("Pre-Prep" in str(ws.cell(r, c).value or "")
                for c in range(7, min(ws.max_column, 60)))), None)
    if header_row is None:
        print(f"  {path}: no period header row", file=sys.stderr)
        return {}

    periods = read_periods(ws, header_row)
    observed: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    profile: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))

    for row in range(header_row + 1, totals_row):
        raw_acct = ws.cell(row, 1).value
        if raw_acct is None:
            continue
        acct = str(int(raw_acct)) if isinstance(raw_acct, (int, float)) else str(raw_acct).strip()
        if not re.fullmatch(r"\d{3,4}", acct):
            continue
        description = str(ws.cell(row, 3).value or "")
        # Subtotal rows restate their details; counting both doubles the shape.
        if re.search(r"subtotal|- total", description, re.I):
            continue
        for col, phase, offset in periods:
            value = ws.cell(row, col).value
            if isinstance(value, (int, float)) and value:
                observed[acct][phase] += float(value)
                profile[acct][offset] += float(value)

    print(f"  {path}: {len(observed)} accounts across {len(periods)} periods "
          f"(offsets {periods[0][2]} to {periods[-1][2]})", file=sys.stderr)
    return {k: {"phases": dict(v), "profile": dict(profile[k])}
            for k, v in observed.items()}


def normalise(observed: dict[str, dict[str, float]]) -> dict[str, Any]:
    """Convert observed dollars into phase shares, dropping noise."""
    out: dict[str, Any] = {}
    for acct, entry in observed.items():
        phases, weeks = entry["phases"], entry["profile"]
        total = sum(phases.values())
        if abs(total) < MIN_ACCOUNT_TOTAL:
            continue
        shares = {p: v / total for p, v in phases.items() if abs(v / total) > MIN_SHARE}
        if not shares:
            continue
        # Renormalise after dropping noise so the shares still sum to 1.
        scale = sum(shares.values())
        week_total = sum(weeks.values()) or 1.0
        out[acct] = {
            "shares": {p: round(v / scale, 4) for p, v in
                       sorted(shares.items(), key=lambda x: -x[1])},
            # The week-by-week curve, indexed off the first shoot week. This is
            # what carries the ramp that phase shares alone flatten out.
            "profile": {str(k): round(v / week_total, 5)
                        for k, v in sorted(weeks.items())
                        if abs(v / week_total) > 0.001},
            "observed_total": round(total, 2),
            "sources": 1,
        }
    return out


def merge(tables: list[dict[str, Any]]) -> dict[str, Any]:
    """Combine productions, weighting each account by the money behind it."""
    pooled: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    pooled_profile: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    weight: dict[str, float] = defaultdict(float)
    sources: dict[str, int] = defaultdict(int)
    for table in tables:
        for acct, entry in table.items():
            money = abs(entry["observed_total"])
            for phase, share in entry["shares"].items():
                pooled[acct][phase] += share * money
            for offset, share in entry.get("profile", {}).items():
                pooled_profile[acct][offset] += share * money
            weight[acct] += money
            sources[acct] += 1
    merged: dict[str, Any] = {}
    for acct, phases in pooled.items():
        total = weight[acct] or 1.0
        merged[acct] = {
            "shares": {p: round(v / total, 4) for p, v in
                       sorted(phases.items(), key=lambda x: -x[1])},
            "profile": {k: round(v / total, 5) for k, v in
                        sorted(pooled_profile[acct].items(), key=lambda x: int(x[0]))},
            "observed_total": round(weight[acct], 2),
            "sources": sources[acct],
        }
    return merged


def roll_up_departments(table: dict[str, Any]) -> dict[str, Any]:
    """A department-level fallback, for accounts never seen before."""
    pooled: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    pooled_profile: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    weight: dict[str, float] = defaultdict(float)
    for acct, entry in table.items():
        dept = acct[:2] + "00"
        money = abs(entry["observed_total"])
        for phase, share in entry["shares"].items():
            pooled[dept][phase] += share * money
        for offset, share in entry.get("profile", {}).items():
            pooled_profile[dept][offset] += share * money
        weight[dept] += money
    return {dept: {"shares": {p: round(v / (weight[dept] or 1.0), 4)
                              for p, v in sorted(phases.items(), key=lambda x: -x[1])},
                   "profile": {k: round(v / (weight[dept] or 1.0), 5) for k, v in
                               sorted(pooled_profile[dept].items(), key=lambda x: int(x[0]))},
                   "observed_total": round(weight[dept], 2)}
            for dept, phases in pooled.items()}


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("cashflows", nargs="+", help="completed cash flow workbooks")
    ap.add_argument("-o", "--out", default="archetypes.json")
    args = ap.parse_args(argv)

    print("harvesting:", file=sys.stderr)
    tables = [normalise(harvest(p)) for p in args.cashflows]
    accounts = merge(tables)
    departments = roll_up_departments(accounts)

    payload = {
        "source_count": len(args.cashflows),
        "sources": args.cashflows,
        "accounts": accounts,
        "departments": departments,
        "_note": ("Phase shares observed in cash flows cut by hand. Learned, not "
                  "assumed — but from a small sample; treat as a prior, not truth."),
    }
    with open(args.out, "w") as fh:
        json.dump(payload, fh, indent=2)

    print(f"\nwrote {args.out}")
    print(f"{len(accounts)} account-level archetypes · "
          f"{len(departments)} department fallbacks")

    print(f"\n{'ACCT':<7}{'PREP':>8}{'SHOOT':>8}{'POST':>8}{'HIATUS':>8}{'MONEY':>13}")
    print("-" * 52)
    ranked = sorted(accounts.items(), key=lambda x: -abs(x[1]["observed_total"]))
    for acct, entry in ranked[:18]:
        s = entry["shares"]
        print(f"{acct:<7}{s.get('prep', 0) * 100:>7.0f}%{s.get('shoot', 0) * 100:>7.0f}%"
              f"{s.get('post', 0) * 100:>7.0f}%{s.get('hiatus', 0) * 100:>7.0f}%"
              f"{entry['observed_total']:>13,.0f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
