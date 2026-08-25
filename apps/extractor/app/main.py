"""
Budget extraction and document generation, as a service.

This is where the domain logic lives. The Node app owns the database and the
web layer; this service owns reading a budget and turning it into the two
documents a production actually needs.

It is Python because that is where the libraries are — `pdfplumber` has no
equivalent in Node, and the extraction logic here is already verified against
real documents: exact reconciliation to a $6,062,000 budget, 97% on daily rate
derivation, 95% on budgeted day costs against the hot cost a production
accountant produced by hand.

Endpoints:
    POST /extract        budget PDF                  -> structured budget JSON
    POST /generate/hotcost   budget JSON + config    -> day sheets (xlsx)
    POST /generate/cashflow  budget JSON + config    -> weekly grid (json)
    POST /learn              completed cash flow(s)  -> spread profiles
    GET  /health
"""

from __future__ import annotations

import io
import json
import tempfile
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse

from .extract_budget import BudgetParser
from .generate_cashflow import Generator, guard_production_type
from .generate_hotcost import (DEFAULT_CONVENTIONS, build_crew,
                               department_fringe_rates, learn_unit_curves,
                               shoot_days, write_day_sheet,
                               write_summary_sheet, write_weekly_sheet)
from .learn_archetypes import harvest, merge, normalise, roll_up_departments
from .overrides import OverrideSet

app = FastAPI(title="cashflow extractor", version="1.0")

MAX_UPLOAD_BYTES = 64 * 1024 * 1024

# Starlette caps one form field at 1MB. An extracted budget is JSON for every
# account in the document: 938KB for a $6M feature, 2.2MB for a $44M one. The
# cap is not a safety margin here — it is a ceiling the second real production
# walked straight through, with a 400 that says nothing about budgets.
MAX_FIELD_BYTES = 64 * 1024 * 1024


async def _fields(request: Request, *required: str) -> dict[str, str]:
    """Read the text fields of a multipart body of any realistic size."""
    form = await request.form(max_part_size=MAX_FIELD_BYTES)
    fields = {key: value for key, value in form.items() if isinstance(value, str)}
    missing = [name for name in required if name not in fields]
    if missing:
        raise HTTPException(422, f"missing form field(s): {', '.join(missing)}")
    return fields


async def _spool(upload: UploadFile, suffix: str) -> Path:
    """Land an upload on disk — the parsers work from paths, not streams."""
    payload = await upload.read()
    if len(payload) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"{upload.filename} exceeds the 64MB limit")
    handle = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    handle.write(payload)
    handle.close()
    return Path(handle.name)


@app.get("/health")
def health() -> dict[str, Any]:
    return {"ok": True, "service": "cashflow-extractor"}


@app.post("/extract")
async def extract(budget: UploadFile = File(...)) -> JSONResponse:
    """Budget PDF in, structured JSON out.

    The response carries production type detection and the inputs_required
    manifest, so the caller knows both what kind of budget this is and the
    short list of things it could not determine on its own.
    """
    path = await _spool(budget, Path(budget.filename or "budget.pdf").suffix or ".pdf")
    try:
        parser = BudgetParser(str(path))
        parser.parse()
        data = parser.to_dict()
    except SystemExit as exc:
        raise HTTPException(422, str(exc)) from exc
    except Exception as exc:                                   # noqa: BLE001
        raise HTTPException(422, f"could not read {budget.filename}: {exc}") from exc
    finally:
        path.unlink(missing_ok=True)

    data["source"]["file"] = budget.filename
    totals = data.get("totals", {})
    coverage = totals.get("extraction_coverage")
    if coverage is not None and coverage < 0.98:
        # The extract is only trustworthy when it accounts for the whole budget.
        data.setdefault("warnings", []).insert(
            0, f"extraction accounts for only {coverage:.1%} of the stated total")
    return JSONResponse(data)


@app.post("/generate/hotcost")
async def generate_hotcost(request: Request) -> StreamingResponse:
    """Pre-populated day sheets. The accountant enters call, lunch and wrap."""
    fields = await _fields(request, "budget_json", "production_json")
    try:
        budget = json.loads(fields["budget_json"])
        cfg = json.loads(fields["production_json"])
    except json.JSONDecodeError as exc:
        raise HTTPException(400, f"invalid JSON: {exc}") from exc

    import openpyxl

    conventions = {**DEFAULT_CONVENTIONS, **cfg.get("hot_cost_conventions", {})}
    curves = learn_unit_curves(budget)
    fringe_rates = department_fringe_rates(budget)
    crew = build_crew(budget)
    included = conventions.get("departments")
    if included:
        crew = [m for m in crew if m["department"] in included]
    if not crew:
        raise HTTPException(422, "no crew found in this budget")

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("SUMMARY")
    weekly = workbook.create_sheet("WEEKLY")

    sheets: list[dict[str, Any]] = []
    for day_no, when, phase in shoot_days(cfg):
        title = (when.strftime("%m%d%y") if phase == "shoot"
                 else when.strftime("%m%d%y") + " PRESHOOT")[:31]
        meta = write_day_sheet(workbook.create_sheet(title), crew, day_no, when,
                               phase, curves, conventions, fringe_rates)
        meta.update({
            "title": title,
            "label": when.strftime("%d %b") if phase == "shoot" else "preshoot",
            "week": 0 if phase != "shoot" else (day_no - 1) // 5 + 1,
        })
        sheets.append(meta)

    write_summary_sheet(summary, sheets)
    write_weekly_sheet(weekly, sheets)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    title = (budget.get("production", {}).get("production_number") or "hotcost")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{title}-hotcost.xlsx"',
                 "X-Crew-Count": str(len(crew)),
                 "X-Day-Sheets": str(len(sheets))})


@app.post("/generate/cashflow")
async def generate_cashflow(request: Request) -> JSONResponse:
    """Weekly grid, reconciled to the budget or not emitted at all."""
    fields = await _fields(request, "budget_json", "production_json")
    force = str(fields.get("force", "")).lower() in ("1", "true", "yes", "on")
    try:
        budget = json.loads(fields["budget_json"])
        cfg = json.loads(fields["production_json"])
        archetypes = json.loads(fields.get("archetypes_json") or "null")
        override_payload = json.loads(fields.get("overrides_json") or "null")
    except json.JSONDecodeError as exc:
        raise HTTPException(400, f"invalid JSON: {exc}") from exc

    try:
        guard_production_type(budget, force)
    except SystemExit as exc:
        # Television is refused rather than silently treated as a feature.
        raise HTTPException(422, str(exc)) from exc

    overrides = OverrideSet()
    if override_payload:
        with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False) as handle:
            json.dump(override_payload, handle)
            temp = handle.name
        overrides = OverrideSet.load(temp)
        Path(temp).unlink(missing_ok=True)

    try:
        result = Generator(budget, cfg, archetypes, overrides).build()
    except SystemExit as exc:
        raise HTTPException(422, str(exc)) from exc

    if not result["reconciliation"]["reconciles"]:
        raise HTTPException(422, {
            "error": "generated schedule does not reconcile to the budget",
            "reconciliation": result["reconciliation"],
        })
    return JSONResponse(result)


@app.post("/learn")
async def learn(cashflows: list[UploadFile] = File(...)) -> JSONResponse:
    """Observed spread profiles from cash flows someone cut by hand.

    Every completed cash flow is a labelled example: it states what each account
    was actually paid, week by week. More productions sharpen every account they
    touch, which a hand-written archetype table cannot do.
    """
    tables, paths = [], []
    try:
        for upload in cashflows:
            path = await _spool(upload, Path(upload.filename or "cf.xlsx").suffix or ".xlsx")
            paths.append(path)
            tables.append(normalise(harvest(str(path))))
    finally:
        for path in paths:
            path.unlink(missing_ok=True)

    accounts = merge(tables)
    if not accounts:
        raise HTTPException(422, "no cash flow grid found in the uploaded file(s)")
    return JSONResponse({
        "source_count": len(cashflows),
        "accounts": accounts,
        "departments": roll_up_departments(accounts),
    })
